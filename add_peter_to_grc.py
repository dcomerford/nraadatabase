"""Add Peter's June 13 formula columns + leaderboard to the GRC master xlsx.

Peter's formula: Adjusted = (Raw Score + Centres) × Equipment Factor
  - NO conversion (no 1.2× for TR/Sporter)
  - NO 0.7 centre weight (centres count fully)

Factors:
  TR   1.53
  FO   1.39
  FS   1.43
  FTR  1.39
  SO   1.47
  SP   1.53

Source: GRC_MCSI_Club_Champion_Master_2026_v5.xlsx (has V4 + V5 already)
Output: GRC_MCSI_Club_Champion_Master_2026_v5_with_peter.xlsx
"""
import shutil
import warnings
warnings.filterwarnings('ignore')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

SRC = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_MCSI_Club_Champion_Master_2026_v5.xlsx'
OUT = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_MCSI_Club_Champion_Master_2026_v5_with_peter.xlsx'

PETER = {'FO': 1.39, 'FTR': 1.39, 'FS': 1.43, 'TR': 1.53, 'SO': 1.47, 'SP': 1.53}


def peter_adjusted(disc, raw, centres):
    """Peter's: (raw + centres) × factor. No conversion, no 0.7 weight."""
    return round((raw + centres) * PETER[disc], 2)


shutil.copyfile(SRC, OUT)
wb = openpyxl.load_workbook(OUT)

NAVY = PatternFill('solid', fgColor='1F4E78')
LITE_RED = PatternFill('solid', fgColor='FCE5E5')
WHITE_BOLD = Font(bold=True, color='FFFFFF')
HEAD = Font(bold=True)
ITAL_GREY = Font(italic=True, color='666666')
CENTER = Alignment(horizontal='center')

# --- Results sheet: add 'Adjusted Peter' column ---
ws = wb['Results']
# find next free col after Adjusted V5 (col 12)
peter_col = None
for col in range(13, 20):
    if ws.cell(1, col).value is None:
        peter_col = col
        break
if peter_col is None:
    peter_col = 13

ws.cell(1, peter_col, 'Adjusted Peter')
ws.cell(1, peter_col).fill = LITE_RED
ws.cell(1, peter_col).font = HEAD
for r in range(2, ws.max_row + 1):
    disc = ws.cell(r, 4).value
    raw = ws.cell(r, 5).value
    cen = ws.cell(r, 6).value
    if disc in PETER and raw is not None and cen is not None:
        ws.cell(r, peter_col, peter_adjusted(disc, raw, int(cen)))
ws.column_dimensions[get_column_letter(peter_col)].width = 14

# --- Build Club Champion Peter leaderboard ---
strings_by_shooter = defaultdict(lambda: defaultdict(list))
disc_by_shooter = {}
for r in range(2, ws.max_row + 1):
    shooter = ws.cell(r, 3).value
    disc = ws.cell(r, 4).value
    dist = ws.cell(r, 2).value
    raw = ws.cell(r, 5).value
    cen = ws.cell(r, 6).value
    if not (shooter and disc in PETER and raw is not None and cen is not None):
        continue
    strings_by_shooter[shooter][dist].append(peter_adjusted(disc, raw, int(cen)))
    # Peter keeps SO and SP separate (different factors)
    disc_by_shooter[shooter] = {'SO':'Sporter O','SP':'Sporter P'}.get(disc, disc)

leaderboard = []
DISTS = (500, 600, 700, 800, 900)
for shooter, by_dist in strings_by_shooter.items():
    total = 0.0; cells = {}; n_dists = 0
    for d in DISTS:
        scores = sorted(by_dist.get(d, []), reverse=True)
        b1 = scores[0] if scores else None
        b2 = scores[1] if len(scores) >= 2 else None
        cells[d] = (b1, b2)
        if b1 is not None: total += b1; n_dists += 1
        if b2 is not None: total += b2
    leaderboard.append({'shooter':shooter, 'disc':disc_by_shooter[shooter],
                        'cells':cells, 'total':round(total,2), 'n':n_dists})
leaderboard.sort(key=lambda r: -r['total'])

if 'Club Champion Peter' in wb.sheetnames:
    del wb['Club Champion Peter']
new = wb.create_sheet('Club Champion Peter')

new.cell(1, 1, "End of Year MCSI Club Champion 2026 — Peter's June 13 formula").font = HEAD
new.cell(2, 1, "Best-2-per-distance using Peter's factors. Formula: Adjusted = (Raw + Centres) × Factor "
               "(no conversion, no 0.7 centre weight). SO and SP kept separate per Peter's "
               "original specification.").font = ITAL_GREY

header = ['Shooter', 'Disc',
          '500m B1', '500m B2', '600m B1', '600m B2',
          '700m B1', '700m B2', '800m B1', '800m B2',
          '900m B1', '900m B2', 'Grand Total', 'Rank']
for col, h in enumerate(header, 1):
    c = new.cell(3, col, h)
    c.fill = NAVY; c.font = WHITE_BOLD; c.alignment = CENTER

for rank, e in enumerate(leaderboard, 1):
    row = 3 + rank
    new.cell(row, 1, e['shooter'])
    new.cell(row, 2, e['disc']).alignment = CENTER
    col = 3
    for d in DISTS:
        b1, b2 = e['cells'][d]
        new.cell(row, col, b1 if b1 is not None else '')
        new.cell(row, col+1, b2 if b2 is not None else '')
        col += 2
    new.cell(row, 13, e['total']).font = HEAD
    new.cell(row, 14, rank).alignment = CENTER
    if rank <= 3:
        for c in range(1, 15): new.cell(row, c).fill = LITE_RED

widths = [26, 10] + [11]*10 + [13, 6]
for i, w in enumerate(widths, 1):
    new.column_dimensions[get_column_letter(i)].width = w
new.freeze_panes = 'C4'

wb.save(OUT)
print(f'Wrote {OUT}')
print(f'  - {len(leaderboard)} shooters in Peter\'s leaderboard')
print(f'\nTop 10 by Peter\'s formula:')
for i, e in enumerate(leaderboard[:10], 1):
    print(f'  {i:>2}. {e["shooter"]:<26} ({e["disc"]:<10}) {e["total"]:>7.2f}')
