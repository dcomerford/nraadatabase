"""Take Adrian's GRC_MCSI_Club_Champion_Master xlsx and add V4-MOD analysis:
- New column in Settings comparison table
- New 'Adjusted V4-MOD' column in Results
- New 'Club Champion V4-MOD' sheet
"""
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

SRC = '/Users/dancomerford/Downloads/GRC_MCSI_Club_Champion_Master_2026_v4.xlsx'
OUT = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_MCSI_Club_Champion_Master_2026_v4_with_V4MOD.xlsx'

# V4-MOD factors
V4MOD = {'FO':1.42, 'FTR':1.45, 'FS':1.44, 'TR':1.42, 'SO':1.42, 'SP':1.43}
# V4 (for comparison column we already see)
V4    = {'FO':1.42, 'FTR':1.45, 'FS':1.46, 'TR':1.42, 'SO':1.40, 'SP':1.41}

# Score conversion: 1.2 for TR/SO/SP, 1.0 for F-class
SC = {'FO':1.0,'FTR':1.0,'FS':1.0,'TR':1.2,'SO':1.2,'SP':1.2}
CENTRE = 0.7

def adjusted(disc, raw, centres, factors):
    eq = raw * SC[disc]
    return round((eq + centres * CENTRE) * factors[disc], 2)


shutil.copyfile(SRC, OUT)
import warnings
warnings.filterwarnings('ignore')
wb = openpyxl.load_workbook(OUT)

NAVY = PatternFill('solid', fgColor='1F4E78')
LITE = PatternFill('solid', fgColor='DDEBF7')
WHITE_BOLD = Font(bold=True, color='FFFFFF')
HEAD = Font(bold=True)
CENTER = Alignment(horizontal='center')

# -------------------------------------------------------------------------
# 1. SETTINGS — add V4-MOD column to the comparison block on the right
# -------------------------------------------------------------------------
ws = wb['Settings']
# Existing comparison block is columns F:J (Discipline, V4, V3, V2, Peter's) in rows 1..7
# Add V4-MOD as new column K (col 11)
v4mod_col = 11
ws.cell(1, v4mod_col, 'V4-MOD').font = HEAD
ws.cell(1, v4mod_col).fill = LITE
disc_to_row = {ws.cell(r, 6).value: r for r in range(2, 8) if ws.cell(r, 6).value}
name_to_short = {
    'F-Open':'FO', 'F/TR':'FTR', 'F-Standard':'FS',
    'Target Rifle':'TR', 'Sporter Open':'SO', 'Sporter Production':'SP',
}
for name, row in disc_to_row.items():
    short = name_to_short.get(name)
    if short:
        ws.cell(row, v4mod_col, V4MOD[short])
        ws.cell(row, v4mod_col).fill = LITE
ws.column_dimensions[get_column_letter(v4mod_col)].width = 10

# -------------------------------------------------------------------------
# 2. RESULTS — add 'Adjusted V4-MOD' column
# -------------------------------------------------------------------------
ws = wb['Results']
# Existing columns: A:K (Date Range Shooter Discipline Raw Centres SC Equalised CentreBonus EF Adjusted)
v4mod_col = 12
ws.cell(1, v4mod_col, 'Adjusted V4-MOD')
ws.cell(1, v4mod_col).fill = LITE
ws.cell(1, v4mod_col).font = HEAD
for r in range(2, ws.max_row + 1):
    disc = ws.cell(r, 4).value
    raw = ws.cell(r, 5).value
    cen = ws.cell(r, 6).value
    if disc in V4MOD and raw is not None and cen is not None:
        ws.cell(r, v4mod_col, adjusted(disc, raw, int(cen), V4MOD))
ws.column_dimensions[get_column_letter(v4mod_col)].width = 15

# -------------------------------------------------------------------------
# 3. NEW SHEET — 'Club Champion V4-MOD' leaderboard
# -------------------------------------------------------------------------
# Pull every result row, group by (shooter, distance), take best 2 V4-MOD adjusted, sum
ws_results = wb['Results']
strings_by_shooter = defaultdict(lambda: defaultdict(list))
disc_by_shooter = {}
for r in range(2, ws_results.max_row + 1):
    shooter = ws_results.cell(r, 3).value
    disc    = ws_results.cell(r, 4).value
    dist    = ws_results.cell(r, 2).value
    raw     = ws_results.cell(r, 5).value
    cen     = ws_results.cell(r, 6).value
    if not (shooter and disc in V4MOD and raw is not None and cen is not None):
        continue
    adj = adjusted(disc, raw, int(cen), V4MOD)
    strings_by_shooter[shooter][dist].append(adj)
    disc_by_shooter[shooter] = disc

# Build leaderboard
leaderboard = []
DISTS = (500, 600, 700, 800, 900)
for shooter, by_dist in strings_by_shooter.items():
    total = 0.0
    cell_data = {}
    n_dists = 0
    for dist in DISTS:
        scores = sorted(by_dist.get(dist, []), reverse=True)
        b1 = scores[0] if len(scores) >= 1 else None
        b2 = scores[1] if len(scores) >= 2 else None
        cell_data[dist] = (b1, b2)
        if b1 is not None: total += b1; n_dists += 1
        if b2 is not None: total += b2
    leaderboard.append({
        'shooter': shooter, 'disc': disc_by_shooter[shooter],
        'cells': cell_data, 'total': round(total, 2), 'n': n_dists,
    })

leaderboard.sort(key=lambda r: -r['total'])

# Create or replace sheet
if 'Club Champion V4-MOD' in wb.sheetnames:
    del wb['Club Champion V4-MOD']
new = wb.create_sheet('Club Champion V4-MOD')

new.cell(1, 1, 'End of Year MCSI Club Champion 2026 — V4-MOD factors').font = HEAD
new.cell(2, 1, "Same best-2-per-distance method as the V4 tab, but using V4-MOD factors "
               "(F-Std 1.46→1.44, SO 1.40→1.42, SP 1.41→1.43).").font = Font(italic=True, color='666666')

header = ['Shooter', 'Disc',
          '500m Best1', '500m Best2',
          '600m Best1', '600m Best2',
          '700m Best1', '700m Best2',
          '800m Best1', '800m Best2',
          '900m Best1', '900m Best2',
          'Grand Total', 'Rank']
for col, h in enumerate(header, 1):
    c = new.cell(3, col, h)
    c.fill = NAVY; c.font = WHITE_BOLD; c.alignment = CENTER

for rank, e in enumerate(leaderboard, 1):
    row = 3 + rank
    new.cell(row, 1, e['shooter'])
    new.cell(row, 2, e['disc']).alignment = CENTER
    col = 3
    for dist in DISTS:
        b1, b2 = e['cells'][dist]
        new.cell(row, col,   b1 if b1 is not None else '')
        new.cell(row, col+1, b2 if b2 is not None else '')
        col += 2
    new.cell(row, 13, e['total']).font = HEAD
    new.cell(row, 14, rank).alignment = CENTER
    if rank <= 3:
        for c in range(1, 15):
            new.cell(row, c).fill = LITE

widths = [26, 6] + [12]*10 + [13, 6]
for i, w in enumerate(widths, 1):
    new.column_dimensions[get_column_letter(i)].width = w
new.freeze_panes = 'C4'

wb.save(OUT)
print(f'Wrote {OUT}')
print(f'  - {len(leaderboard)} shooters in V4-MOD leaderboard')
print(f'  - Settings now includes V4-MOD column')
print(f'  - Results now includes Adjusted V4-MOD column')
print(f'\nTop 5 by V4-MOD:')
for i, e in enumerate(leaderboard[:5], 1):
    print(f'  {i}. {e["shooter"]:<24} ({e["disc"]:<3}) {e["total"]:.2f}')
