"""Add V5.1 factors + columns + leaderboard sheet to the GRC master xlsx.

V5.1 = V4 structure (Score × Conversion + Centres × 0.7) × Factor.

V5.1 factors (calibrated on K&Q 2024+ AND 31 same-conditions prize meets,
top-40% per bucket, MIN_BUCKET_N=10):
  TR     1.403  (V5: 1.412)
  F-Open 1.404  (V5: 1.406)
  F-Std  1.484  (V5: 1.475)
  FTR    1.444  (V5: 1.450)
  SO     1.325  (V5: 1.383)  ← Sporter dropped 0.058 with broader sample
  SP     1.325  (V5: 1.383)

Source: GRC_MCSI_Club_Champion_Master_2026_v5.xlsx (which already has V4 + V5).
Output: GRC_MCSI_Club_Champion_Master_2026_v5.1.xlsx (adds V5.1 columns + sheet).
"""
import shutil
import warnings
warnings.filterwarnings('ignore')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

SRC = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_MCSI_Club_Champion_Master_2026_v5.xlsx'
OUT = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_MCSI_Club_Champion_Master_2026_v5.1.xlsx'

V51 = {'FO': 1.404, 'FTR': 1.444, 'FS': 1.484, 'TR': 1.403, 'SO': 1.325, 'SP': 1.325}
SC  = {'FO': 1.0,   'FTR': 1.0,   'FS': 1.0,   'TR': 1.2,   'SO': 1.2,   'SP': 1.2}
CENTRE = 0.7


def adjusted(disc, raw, centres):
    eq = raw * SC[disc]
    return round((eq + centres * CENTRE) * V51[disc], 2)


shutil.copyfile(SRC, OUT)
wb = openpyxl.load_workbook(OUT)

NAVY = PatternFill('solid', fgColor='1F4E78')
LITE_AMBER = PatternFill('solid', fgColor='FFF2CC')  # V5.1 = amber to distinguish from V5 green
WHITE_BOLD = Font(bold=True, color='FFFFFF')
HEAD = Font(bold=True)
ITAL_GREY = Font(italic=True, color='666666')
CENTER = Alignment(horizontal='center')

# ---------------------------------------------------------------------------
# 1. SETTINGS — find next free column after the existing comparison block
# ---------------------------------------------------------------------------
ws = wb['Settings']
# Probe: scan row 1 to find first empty col after V5 (which add_v5_to_grc put in col 11)
v51_col = None
for col in range(12, 20):
    if ws.cell(1, col).value is None:
        v51_col = col
        break
if v51_col is None:
    v51_col = 12

ws.cell(1, v51_col, 'V5.1').font = HEAD
ws.cell(1, v51_col).fill = LITE_AMBER
disc_to_row = {ws.cell(r, 6).value: r for r in range(2, 8) if ws.cell(r, 6).value}
name_to_short = {
    'F-Open': 'FO', 'F/TR': 'FTR', 'F-Standard': 'FS',
    'Target Rifle': 'TR', 'Sporter Open': 'SO', 'Sporter Production': 'SP',
}
for name, row in disc_to_row.items():
    short = name_to_short.get(name)
    if short:
        ws.cell(row, v51_col, V51[short])
        ws.cell(row, v51_col).fill = LITE_AMBER
ws.column_dimensions[get_column_letter(v51_col)].width = 10

# Notes block — add V5.1 note below the existing V5 note
note_row = 12
ws.cell(note_row, 6, "V5.1: Recalibrated on K&Q + 31 same-conditions prize meets "
                     "(28,213 strings). Sporter factor dropped to 1.325 with broader sample. "
                     "MIN_BUCKET_N=10. Sporter still merged.").font = ITAL_GREY

# ---------------------------------------------------------------------------
# 2. RESULTS — add 'Adjusted V5.1' column after 'Adjusted V5'
# ---------------------------------------------------------------------------
ws = wb['Results']
# Find next free col after Adjusted V5 (which add_v5_to_grc put in col 12)
v51_res_col = None
for col in range(13, 20):
    if ws.cell(1, col).value is None:
        v51_res_col = col
        break
if v51_res_col is None:
    v51_res_col = 13

ws.cell(1, v51_res_col, 'Adjusted V5.1')
ws.cell(1, v51_res_col).fill = LITE_AMBER
ws.cell(1, v51_res_col).font = HEAD
for r in range(2, ws.max_row + 1):
    disc = ws.cell(r, 4).value
    raw = ws.cell(r, 5).value
    cen = ws.cell(r, 6).value
    if disc in V51 and raw is not None and cen is not None:
        ws.cell(r, v51_res_col, adjusted(disc, raw, int(cen)))
ws.column_dimensions[get_column_letter(v51_res_col)].width = 14

# ---------------------------------------------------------------------------
# 3. NEW SHEET — Club Champion V5.1 leaderboard
# ---------------------------------------------------------------------------
ws_results = wb['Results']
strings_by_shooter = defaultdict(lambda: defaultdict(list))
disc_by_shooter = {}
for r in range(2, ws_results.max_row + 1):
    shooter = ws_results.cell(r, 3).value
    disc = ws_results.cell(r, 4).value
    dist = ws_results.cell(r, 2).value
    raw = ws_results.cell(r, 5).value
    cen = ws_results.cell(r, 6).value
    if not (shooter and disc in V51 and raw is not None and cen is not None):
        continue
    adj = adjusted(disc, raw, int(cen))
    strings_by_shooter[shooter][dist].append(adj)
    disc_by_shooter[shooter] = 'Sporter' if disc in ('SO', 'SP') else disc

leaderboard = []
DISTS = (500, 600, 700, 800, 900)
for shooter, by_dist in strings_by_shooter.items():
    total = 0.0
    cell_data = {}
    n_dists = 0
    for dist in DISTS:
        scores = sorted(by_dist.get(dist, []), reverse=True)
        b1 = scores[0] if len(scores) >= 1 else None
        b2 = scores[1] if len(scores) >= 2 else None
        cell_data[dist] = (b1, b2)
        if b1 is not None: total += b1; n_dists += 1
        if b2 is not None: total += b2
    leaderboard.append({
        'shooter': shooter, 'disc': disc_by_shooter[shooter],
        'cells': cell_data, 'total': round(total, 2), 'n': n_dists,
    })

leaderboard.sort(key=lambda r: -r['total'])

if 'Club Champion V5.1' in wb.sheetnames:
    del wb['Club Champion V5.1']
new = wb.create_sheet('Club Champion V5.1')

new.cell(1, 1, 'End of Year MCSI Club Champion 2026 — V5.1 factors').font = HEAD
new.cell(2, 1, "Best-2-per-distance using V5.1 factors. "
               "V5.1 calibrated on K&Q 2024+ AND 31 same-conditions prize meets "
               "(28,213 strings). Sporter Open + Sporter Production still merged.").font = ITAL_GREY

header = ['Shooter', 'Disc',
          '500m Best 1', '500m Best 2',
          '600m Best 1', '600m Best 2',
          '700m Best 1', '700m Best 2',
          '800m Best 1', '800m Best 2',
          '900m Best 1', '900m Best 2',
          'Grand Total', 'Rank']
for col, h in enumerate(header, 1):
    c = new.cell(3, col, h)
    c.fill = NAVY; c.font = WHITE_BOLD; c.alignment = CENTER

for rank, e in enumerate(leaderboard, 1):
    row = 3 + rank
    new.cell(row, 1, e['shooter'])
    new.cell(row, 2, e['disc']).alignment = CENTER
    col = 3
    for dist in DISTS:
        b1, b2 = e['cells'][dist]
        new.cell(row, col, b1 if b1 is not None else '')
        new.cell(row, col + 1, b2 if b2 is not None else '')
        col += 2
    new.cell(row, 13, e['total']).font = HEAD
    new.cell(row, 14, rank).alignment = CENTER
    if rank <= 3:
        for c in range(1, 15):
            new.cell(row, c).fill = LITE_AMBER

widths = [26, 9] + [12]*10 + [13, 6]
for i, w in enumerate(widths, 1):
    new.column_dimensions[get_column_letter(i)].width = w
new.freeze_panes = 'C4'

wb.save(OUT)
print(f'Wrote {OUT}')
print(f'  - {len(leaderboard)} shooters in V5.1 leaderboard')
print(f'  - Settings now includes V4, V5, V5.1 factor columns')
print(f'  - Results now includes Adjusted V4, V5, V5.1 columns')
print(f'\nTop 10 by V5.1:')
for i, e in enumerate(leaderboard[:10], 1):
    print(f'  {i:>2}. {e["shooter"]:<26} ({e["disc"]:<8}) {e["total"]:>7.2f}')
