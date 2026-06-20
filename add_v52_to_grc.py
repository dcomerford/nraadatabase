"""Replace V5.1 with V5.2 in the GRC master xlsx.

V5.2 = K&Q 2024+ East+National only (WARA excluded). See v52_calibration.py.

Factors:
  TR     1.414  (V5: 1.412)
  F-Open 1.410  (V5: 1.406)
  F-Std  1.472  (V5: 1.475)
  FTR    1.448  (V5: 1.450)
  SO/SP  1.366  (V5: 1.383)  ← only material change

Output: GRC_MCSI_Club_Champion_Master_2026_v5.2.xlsx
Removes the V5.1 sheet/column (it was WA-biased and rejected as unfair).
"""
import shutil
import warnings
warnings.filterwarnings('ignore')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

SRC = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_MCSI_Club_Champion_Master_2026_v5.xlsx'  # clean base (V5)
OUT = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_MCSI_Club_Champion_Master_2026_v5.2.xlsx'

V52 = {'FO': 1.410, 'FTR': 1.448, 'FS': 1.472, 'TR': 1.414, 'SO': 1.366, 'SP': 1.366}
SC  = {'FO': 1.0,   'FTR': 1.0,   'FS': 1.0,   'TR': 1.2,   'SO': 1.2,   'SP': 1.2}
CENTRE = 0.7


def adjusted(disc, raw, centres):
    eq = raw * SC[disc]
    return round((eq + centres * CENTRE) * V52[disc], 2)


shutil.copyfile(SRC, OUT)
wb = openpyxl.load_workbook(OUT)

NAVY = PatternFill('solid', fgColor='1F4E78')
LITE_BLUE = PatternFill('solid', fgColor='DDEBF7')
LITE_PURPLE = PatternFill('solid', fgColor='E4D7F5')  # V5.2 = purple to distinguish
WHITE_BOLD = Font(bold=True, color='FFFFFF')
HEAD = Font(bold=True)
ITAL_GREY = Font(italic=True, color='666666')
CENTER = Alignment(horizontal='center')

# ---------------------------------------------------------------------------
# 1. SETTINGS — add V5.2 column after V5
# ---------------------------------------------------------------------------
ws = wb['Settings']
v52_col = None
for col in range(12, 20):
    if ws.cell(1, col).value is None:
        v52_col = col
        break
if v52_col is None:
    v52_col = 12

ws.cell(1, v52_col, 'V5.2').font = HEAD
ws.cell(1, v52_col).fill = LITE_PURPLE
disc_to_row = {ws.cell(r, 6).value: r for r in range(2, 8) if ws.cell(r, 6).value}
name_to_short = {
    'F-Open': 'FO', 'F/TR': 'FTR', 'F-Standard': 'FS',
    'Target Rifle': 'TR', 'Sporter Open': 'SO', 'Sporter Production': 'SP',
}
for name, row in disc_to_row.items():
    short = name_to_short.get(name)
    if short:
        ws.cell(row, v52_col, V52[short])
        ws.cell(row, v52_col).fill = LITE_PURPLE
ws.column_dimensions[get_column_letter(v52_col)].width = 10

note_row = 12
ws.cell(note_row, 6, "V5.2: Recalibrated on K&Q 2024+ East+National only (WARA excluded). "
                     "Prize-meet data was 99% WA, biasing Sporter unfairly. "
                     "V5.2 keeps Sporter merge but lowers factor 0.017 from V5.").font = ITAL_GREY

# ---------------------------------------------------------------------------
# 2. RESULTS — add 'Adjusted V5.2' column
# ---------------------------------------------------------------------------
ws = wb['Results']
v52_res_col = None
for col in range(13, 20):
    if ws.cell(1, col).value is None:
        v52_res_col = col
        break
if v52_res_col is None:
    v52_res_col = 13

ws.cell(1, v52_res_col, 'Adjusted V5.2')
ws.cell(1, v52_res_col).fill = LITE_PURPLE
ws.cell(1, v52_res_col).font = HEAD
for r in range(2, ws.max_row + 1):
    disc = ws.cell(r, 4).value
    raw = ws.cell(r, 5).value
    cen = ws.cell(r, 6).value
    if disc in V52 and raw is not None and cen is not None:
        ws.cell(r, v52_res_col, adjusted(disc, raw, int(cen)))
ws.column_dimensions[get_column_letter(v52_res_col)].width = 14

# ---------------------------------------------------------------------------
# 3. NEW SHEET — Club Champion V5.2 leaderboard
# ---------------------------------------------------------------------------
ws_results = wb['Results']
strings_by_shooter = defaultdict(lambda: defaultdict(list))
disc_by_shooter = {}
for r in range(2, ws_results.max_row + 1):
    shooter = ws_results.cell(r, 3).value
    disc = ws_results.cell(r, 4).value
    dist = ws_results.cell(r, 2).value
    raw = ws_results.cell(r, 5).value
    cen = ws_results.cell(r, 6).value
    if not (shooter and disc in V52 and raw is not None and cen is not None):
        continue
    adj = adjusted(disc, raw, int(cen))
    strings_by_shooter[shooter][dist].append(adj)
    disc_by_shooter[shooter] = 'Sporter' if disc in ('SO', 'SP') else disc

leaderboard = []
DISTS = (500, 600, 700, 800, 900)
for shooter, by_dist in strings_by_shooter.items():
    total = 0.0
    cell_data = {}
    n_dists = 0
    for dist in DISTS:
        scores = sorted(by_dist.get(dist, []), reverse=True)
        b1 = scores[0] if len(scores) >= 1 else None
        b2 = scores[1] if len(scores) >= 2 else None
        cell_data[dist] = (b1, b2)
        if b1 is not None: total += b1; n_dists += 1
        if b2 is not None: total += b2
    leaderboard.append({
        'shooter': shooter, 'disc': disc_by_shooter[shooter],
        'cells': cell_data, 'total': round(total, 2), 'n': n_dists,
    })

leaderboard.sort(key=lambda r: -r['total'])

if 'Club Champion V5.2' in wb.sheetnames:
    del wb['Club Champion V5.2']
new = wb.create_sheet('Club Champion V5.2')

new.cell(1, 1, 'End of Year MCSI Club Champion 2026 — V5.2 factors').font = HEAD
new.cell(2, 1, "Best-2-per-distance using V5.2 factors. Calibrated on K&Q 2024+ "
               "East+National only (WARA excluded). Sporter Open + Production merged. "
               "Small refinement over V5: Sporter -0.017, others within ±0.004.").font = ITAL_GREY

header = ['Shooter', 'Disc',
          '500m Best 1', '500m Best 2',
          '600m Best 1', '600m Best 2',
          '700m Best 1', '700m Best 2',
          '800m Best 1', '800m Best 2',
          '900m Best 1', '900m Best 2',
          'Grand Total', 'Rank']
for col, h in enumerate(header, 1):
    c = new.cell(3, col, h)
    c.fill = NAVY; c.font = WHITE_BOLD; c.alignment = CENTER

for rank, e in enumerate(leaderboard, 1):
    row = 3 + rank
    new.cell(row, 1, e['shooter'])
    new.cell(row, 2, e['disc']).alignment = CENTER
    col = 3
    for dist in DISTS:
        b1, b2 = e['cells'][dist]
        new.cell(row, col, b1 if b1 is not None else '')
        new.cell(row, col + 1, b2 if b2 is not None else '')
        col += 2
    new.cell(row, 13, e['total']).font = HEAD
    new.cell(row, 14, rank).alignment = CENTER
    if rank <= 3:
        for c in range(1, 15):
            new.cell(row, c).fill = LITE_PURPLE

widths = [26, 9] + [12]*10 + [13, 6]
for i, w in enumerate(widths, 1):
    new.column_dimensions[get_column_letter(i)].width = w
new.freeze_panes = 'C4'

wb.save(OUT)
print(f'Wrote {OUT}')
print(f'  - {len(leaderboard)} shooters in V5.2 leaderboard')
print(f'  - Settings includes V4, V5, V5.2 factor columns')
print(f'  - Results includes Adjusted V4, V5, V5.2 columns')
print(f'\nTop 10 by V5.2:')
for i, e in enumerate(leaderboard[:10], 1):
    print(f'  {i:>2}. {e["shooter"]:<26} ({e["disc"]:<8}) {e["total"]:>7.2f}')
