"""Add V5.5 (middle-20%, K&Q + 99 East PMs, no WA) to the GRC master xlsx."""
import shutil
import warnings
warnings.filterwarnings('ignore')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

SRC = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_MCSI_Club_Champion_Master_2026_v5.xlsx'
OUT = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_MCSI_Club_Champion_Master_2026_v5.5.xlsx'

V55 = {'FO': 1.413, 'FTR': 1.427, 'FS': 1.492, 'TR': 1.362, 'SO': 1.317, 'SP': 1.317}
SC  = {'FO': 1.0,   'FTR': 1.0,   'FS': 1.0,   'TR': 1.2,   'SO': 1.2,   'SP': 1.2}
CENTRE = 0.7


def adjusted(disc, raw, centres):
    eq = raw * SC[disc]
    return round((eq + centres * CENTRE) * V55[disc], 2)


shutil.copyfile(SRC, OUT)
wb = openpyxl.load_workbook(OUT)

NAVY = PatternFill('solid', fgColor='1F4E78')
LITE_ORANGE = PatternFill('solid', fgColor='FCE4D6')
WHITE_BOLD = Font(bold=True, color='FFFFFF')
HEAD = Font(bold=True)
ITAL_GREY = Font(italic=True, color='666666')
CENTER = Alignment(horizontal='center')

# Settings
ws = wb['Settings']
v55_col = None
for col in range(12, 20):
    if ws.cell(1, col).value is None:
        v55_col = col; break
if v55_col is None: v55_col = 12
ws.cell(1, v55_col, 'V5.5').font = HEAD
ws.cell(1, v55_col).fill = LITE_ORANGE
disc_to_row = {ws.cell(r, 6).value: r for r in range(2, 8) if ws.cell(r, 6).value}
name_to_short = {'F-Open':'FO','F/TR':'FTR','F-Standard':'FS','Target Rifle':'TR',
                 'Sporter Open':'SO','Sporter Production':'SP'}
for name, row in disc_to_row.items():
    short = name_to_short.get(name)
    if short:
        ws.cell(row, v55_col, V55[short])
        ws.cell(row, v55_col).fill = LITE_ORANGE
ws.column_dimensions[get_column_letter(v55_col)].width = 10
ws.cell(12, 6, "V5.5: middle-20% calibration on K&Q + 99 East prize meets all-5 (42,232 strings). "
               "Reflects club-level participation. Sporter -0.066, TR -0.050.").font = ITAL_GREY

# Results
ws = wb['Results']
v55_res_col = None
for col in range(13, 20):
    if ws.cell(1, col).value is None:
        v55_res_col = col; break
if v55_res_col is None: v55_res_col = 13
ws.cell(1, v55_res_col, 'Adjusted V5.5')
ws.cell(1, v55_res_col).fill = LITE_ORANGE
ws.cell(1, v55_res_col).font = HEAD
for r in range(2, ws.max_row + 1):
    disc = ws.cell(r, 4).value
    raw = ws.cell(r, 5).value
    cen = ws.cell(r, 6).value
    if disc in V55 and raw is not None and cen is not None:
        ws.cell(r, v55_res_col, adjusted(disc, raw, int(cen)))
ws.column_dimensions[get_column_letter(v55_res_col)].width = 14

# Leaderboard
ws_results = wb['Results']
strings_by_shooter = defaultdict(lambda: defaultdict(list))
disc_by_shooter = {}
for r in range(2, ws_results.max_row + 1):
    shooter = ws_results.cell(r, 3).value
    disc = ws_results.cell(r, 4).value
    dist = ws_results.cell(r, 2).value
    raw = ws_results.cell(r, 5).value
    cen = ws_results.cell(r, 6).value
    if not (shooter and disc in V55 and raw is not None and cen is not None): continue
    strings_by_shooter[shooter][dist].append(adjusted(disc, raw, int(cen)))
    disc_by_shooter[shooter] = 'Sporter' if disc in ('SO','SP') else disc

leaderboard = []
DISTS = (500, 600, 700, 800, 900)
for shooter, by_dist in strings_by_shooter.items():
    total = 0.0; cells = {}; n_dists = 0
    for d in DISTS:
        scores = sorted(by_dist.get(d, []), reverse=True)
        b1 = scores[0] if scores else None
        b2 = scores[1] if len(scores) >= 2 else None
        cells[d] = (b1, b2)
        if b1 is not None: total += b1; n_dists += 1
        if b2 is not None: total += b2
    leaderboard.append({'shooter':shooter,'disc':disc_by_shooter[shooter],
                        'cells':cells,'total':round(total,2),'n':n_dists})
leaderboard.sort(key=lambda r: -r['total'])

if 'Club Champion V5.5' in wb.sheetnames: del wb['Club Champion V5.5']
new = wb.create_sheet('Club Champion V5.5')
new.cell(1,1,'End of Year MCSI Club Champion 2026 — V5.5 factors (middle-20%)').font = HEAD
new.cell(2,1,"K&Q East+Nat all-5 + 99 East-coast prize meets all-5. NO WA. "
             "Middle-20% calibration. Calibrated against club-level participation patterns.").font = ITAL_GREY

header = ['Shooter','Disc','500 B1','500 B2','600 B1','600 B2','700 B1','700 B2',
          '800 B1','800 B2','900 B1','900 B2','Total','Rank']
for col, h in enumerate(header, 1):
    c = new.cell(3, col, h); c.fill = NAVY; c.font = WHITE_BOLD; c.alignment = CENTER

for rank, e in enumerate(leaderboard, 1):
    row = 3 + rank
    new.cell(row,1,e['shooter']); new.cell(row,2,e['disc']).alignment = CENTER
    col = 3
    for d in DISTS:
        b1, b2 = e['cells'][d]
        new.cell(row, col, b1 if b1 is not None else '')
        new.cell(row, col+1, b2 if b2 is not None else '')
        col += 2
    new.cell(row,13,e['total']).font = HEAD
    new.cell(row,14,rank).alignment = CENTER
    if rank <= 3:
        for c in range(1, 15): new.cell(row, c).fill = LITE_ORANGE

widths = [26,9] + [11]*10 + [13,6]
for i, w in enumerate(widths, 1): new.column_dimensions[get_column_letter(i)].width = w
new.freeze_panes = 'C4'

wb.save(OUT)
print(f'Wrote {OUT}')
print(f'\nTop 15 by V5.5 (middle-20% K&Q + East PMs):')
print(f'{"Rank":<5}{"Shooter":<26}{"Disc":<10}{"V5.5":>10}')
for i, e in enumerate(leaderboard[:15], 1):
    print(f'{i:<5}{e["shooter"]:<26}{e["disc"]:<10}{e["total"]:>10.2f}')
