"""Add V5 factors + columns + leaderboard sheet to Adrian's master GRC xlsx.

V5 = V4 structure (Score × Conversion + Centres × 0.7) × Factor, with:
  TR     1.412
  F-Open 1.406
  F-Std  1.475
  FTR    1.450
  SO     1.383  ← Sporter Open and Sporter Production share one factor in V5
  SP     1.383

Calibrated at top-40% K&Q cohort with Sporter merged (see v5_calibration.py).
"""
import shutil
import warnings
warnings.filterwarnings('ignore')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

SRC = '/Users/dancomerford/Downloads/GRC_MCSI_Club_Champion_Master_2026_v4.xlsx'
OUT = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_MCSI_Club_Champion_Master_2026_v5.xlsx'

V5 = {'FO': 1.406, 'FTR': 1.450, 'FS': 1.475, 'TR': 1.412, 'SO': 1.383, 'SP': 1.383}
SC = {'FO': 1.0, 'FTR': 1.0, 'FS': 1.0, 'TR': 1.2, 'SO': 1.2, 'SP': 1.2}
CENTRE = 0.7


def adjusted(disc, raw, centres):
    eq = raw * SC[disc]
    return round((eq + centres * CENTRE) * V5[disc], 2)


shutil.copyfile(SRC, OUT)
wb = openpyxl.load_workbook(OUT)

NAVY = PatternFill('solid', fgColor='1F4E78')
LITE = PatternFill('solid', fgColor='DDEBF7')
LITE_GREEN = PatternFill('solid', fgColor='E2EFDA')
WHITE_BOLD = Font(bold=True, color='FFFFFF')
HEAD = Font(bold=True)
ITAL_GREY = Font(italic=True, color='666666')
CENTER = Alignment(horizontal='center')

# ---------------------------------------------------------------------------
# 1. SETTINGS — add V5 column to the comparison block
# ---------------------------------------------------------------------------
ws = wb['Settings']
# Existing comparison columns F:J = Discipline, V4, V3, V2, Peter's
# Insert V5 at col K (col 11). The 'Peter ratio' column L stays where it is.
v5_col = 11

# Shift the existing col K (the Peter ratio with no header) one to the right if there's a value
# Actually let's just put V5 in col K and let any pre-existing K shift to L; but simpler: K was None
# in row 1 and held a ratio in rows 2-7 (col 12 in our inspection). Let's check by reading first.
existing_k = [ws.cell(r, v5_col).value for r in range(1, 8)]
if any(v is not None for v in existing_k):
    # shift col K → L (one right) for rows 1-8
    for r in range(1, 9):
        val = ws.cell(r, v5_col).value
        if val is not None:
            ws.cell(r, v5_col + 1, val)
            ws.cell(r, v5_col).value = None

ws.cell(1, v5_col, 'V5').font = HEAD
ws.cell(1, v5_col).fill = LITE_GREEN
disc_to_row = {ws.cell(r, 6).value: r for r in range(2, 8) if ws.cell(r, 6).value}
name_to_short = {
    'F-Open': 'FO', 'F/TR': 'FTR', 'F-Standard': 'FS',
    'Target Rifle': 'TR', 'Sporter Open': 'SO', 'Sporter Production': 'SP',
}
for name, row in disc_to_row.items():
    short = name_to_short.get(name)
    if short:
        ws.cell(row, v5_col, V5[short])
        ws.cell(row, v5_col).fill = LITE_GREEN
ws.column_dimensions[get_column_letter(v5_col)].width = 10

# Note about Sporter merge in V5
note_row = 11
ws.cell(note_row, 6, "V5: SO and SP share a single Sporter factor (1.383). "
                     "Calibrated at top-40% K&Q cohort with Sporter merged.").font = ITAL_GREY

# ---------------------------------------------------------------------------
# 2. RESULTS — add 'Adjusted V5' column
# ---------------------------------------------------------------------------
ws = wb['Results']
v5_col = 12
ws.cell(1, v5_col, 'Adjusted V5')
ws.cell(1, v5_col).fill = LITE_GREEN
ws.cell(1, v5_col).font = HEAD
for r in range(2, ws.max_row + 1):
    disc = ws.cell(r, 4).value
    raw = ws.cell(r, 5).value
    cen = ws.cell(r, 6).value
    if disc in V5 and raw is not None and cen is not None:
        ws.cell(r, v5_col, adjusted(disc, raw, int(cen)))
ws.column_dimensions[get_column_letter(v5_col)].width = 14

# ---------------------------------------------------------------------------
# 3. NEW SHEET — Club Champion V5 leaderboard
# ---------------------------------------------------------------------------
ws_results = wb['Results']
strings_by_shooter = defaultdict(lambda: defaultdict(list))
disc_by_shooter = {}
for r in range(2, ws_results.max_row + 1):
    shooter = ws_results.cell(r, 3).value
    disc = ws_results.cell(r, 4).value
    dist = ws_results.cell(r, 2).value
    raw = ws_results.cell(r, 5).value
    cen = ws_results.cell(r, 6).value
    if not (shooter and disc in V5 and raw is not None and cen is not None):
        continue
    adj = adjusted(disc, raw, int(cen))
    strings_by_shooter[shooter][dist].append(adj)
    # Use 'Sporter' as display disc for SO/SP in V5
    disc_by_shooter[shooter] = 'Sporter' if disc in ('SO', 'SP') else disc

leaderboard = []
DISTS = (500, 600, 700, 800, 900)
for shooter, by_dist in strings_by_shooter.items():
    total = 0.0
    cell_data = {}
    n_dists = 0
    for dist in DISTS:
        scores = sorted(by_dist.get(dist, []), reverse=True)
        b1 = scores[0] if len(scores) >= 1 else None
        b2 = scores[1] if len(scores) >= 2 else None
        cell_data[dist] = (b1, b2)
        if b1 is not None: total += b1; n_dists += 1
        if b2 is not None: total += b2
    leaderboard.append({
        'shooter': shooter, 'disc': disc_by_shooter[shooter],
        'cells': cell_data, 'total': round(total, 2), 'n': n_dists,
    })

leaderboard.sort(key=lambda r: -r['total'])

if 'Club Champion V5' in wb.sheetnames:
    del wb['Club Champion V5']
new = wb.create_sheet('Club Champion V5')

new.cell(1, 1, 'End of Year MCSI Club Champion 2026 — V5 factors (Sporter merged)').font = HEAD
new.cell(2, 1, "Best-2-per-distance using V5 factors. "
               "V5 calibrated at top-40% K&Q 2024+ cohort with Sporter Open + "
               "Sporter Production merged into a single Sporter factor (1.383).").font = ITAL_GREY

header = ['Shooter', 'Disc',
          '500m Best 1', '500m Best 2',
          '600m Best 1', '600m Best 2',
          '700m Best 1', '700m Best 2',
          '800m Best 1', '800m Best 2',
          '900m Best 1', '900m Best 2',
          'Grand Total', 'Rank']
for col, h in enumerate(header, 1):
    c = new.cell(3, col, h)
    c.fill = NAVY; c.font = WHITE_BOLD; c.alignment = CENTER

for rank, e in enumerate(leaderboard, 1):
    row = 3 + rank
    new.cell(row, 1, e['shooter'])
    new.cell(row, 2, e['disc']).alignment = CENTER
    col = 3
    for dist in DISTS:
        b1, b2 = e['cells'][dist]
        new.cell(row, col, b1 if b1 is not None else '')
        new.cell(row, col + 1, b2 if b2 is not None else '')
        col += 2
    new.cell(row, 13, e['total']).font = HEAD
    new.cell(row, 14, rank).alignment = CENTER
    if rank <= 3:
        for c in range(1, 15):
            new.cell(row, c).fill = LITE_GREEN

widths = [26, 9] + [12]*10 + [13, 6]
for i, w in enumerate(widths, 1):
    new.column_dimensions[get_column_letter(i)].width = w
new.freeze_panes = 'C4'

wb.save(OUT)
print(f'Wrote {OUT}')
print(f'  - {len(leaderboard)} shooters in V5 leaderboard')
print(f'  - Settings now includes V5 column')
print(f'  - Results now includes Adjusted V5 column')
print(f'\nTop 10 by V5:')
for i, e in enumerate(leaderboard[:10], 1):
    print(f'  {i:>2}. {e["shooter"]:<26} ({e["disc"]:<8}) {e["total"]:>7.2f}')
