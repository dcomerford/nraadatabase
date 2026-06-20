"""GRC Club Championship leaderboard under THREE formulas side-by-side:
  v1   = formula used at June 13 Geelong shoot
         TR 1.45, SP 1.43, SO 1.43, FO 1.44, FS 1.46, FTR 1.46
  v3   = Adrian's final/current set
         TR 1.42, SP 1.41, SO 1.40, FO 1.43, FS 1.46, FTR 1.46
  Pres = President's proposed set
         TR 1.56, SP 1.55, SO 1.46, FO 1.28, FS 1.42, FTR 1.42

Shared: conversion ×1.2 for TR/Sporter, centre value 0.7
Aggregation: best 2 scores per distance summed per shooter

Output: GRC_Club_Championship_3formulas.xlsx
  Tab 1: Leaderboard  — shooter rows, columns for each formula's total + rank
  Tab 2: Master Diffs — rank movement + total deltas between the 3 formulas
"""
import json
import urllib.request
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COMPS = [
    ('0adcba04-e8c4-47a7-9534-fdc73b52d392', 'Jan 26'),
    ('d6d4e0e5-618d-41ad-8dda-b33f7d2f1d39', 'Apr 11'),
    ('d2a4e092-a9cd-47ab-8997-593499363e17', 'May 2'),
    ('c5dee970-b934-4a0b-88ed-44f61e5986c5', 'Jun 13'),
]

FORMULAS = {
    # v1 = original formula (FTR=1.43 assumed — matches FS, not user-confirmed)
    'v1':   {'TR':1.53, 'SP':1.53, 'SO':1.47, 'FO':1.39, 'FS':1.43, 'FTR':1.43},
    'v3':   {'TR':1.42, 'SP':1.41, 'SO':1.40, 'FO':1.43, 'FS':1.46, 'FTR':1.46},
    'Pres': {'TR':1.56, 'SP':1.55, 'SO':1.46, 'FO':1.28, 'FS':1.42, 'FTR':1.42},
}
CENTRE = 0.7

NAME_FIXES = {'Steve Gywn':'Stephen Gwyn', 'Nick Pato':'Nick Patto'}

def to_short(dn):
    if not dn: return None
    d = dn.lower()
    if 'target' in d and 'rifle' in d: return 'TR'
    if 'division open' in d: return 'TR'
    if 'f-tr' in d or 'ftr' in d: return 'FTR'
    if 'f-open' in d or 'fopen' in d or 'f open' in d: return 'FO'
    if 'f-standard' in d or 'f standard' in d or 'fstandard' in d: return 'FS'
    if 'production' in d: return 'SP'
    if 'sporter open' in d: return 'SO'
    if d.strip() == 'sporter': return 'SP'
    return None

def adjust(disc, raw, cen, factors):
    f = factors.get(disc)
    if f is None: return None
    eq = raw if disc in ('FO','FS','FTR') else raw * 1.2
    return round((eq + cen * CENTRE) * f, 3)

def fetch(cid):
    return json.load(urllib.request.urlopen(
        f'https://bulletimpacts.com/api/competitions/{cid}/results'))

# ---- pull all strings ----
strings = []
for cid, label in COMPS:
    d = fetch(cid)
    for rr in d['rangeResults']:
        dist = rr['name']
        for grp in rr['results']:
            short = to_short(grp.get('disciplineName'))
            for e in grp.get('entries', []):
                name = NAME_FIXES.get(
                    f"{e.get('firstName','').strip()} {e.get('lastName','').strip()}".strip(),
                    f"{e.get('firstName','').strip()} {e.get('lastName','').strip()}".strip())
                raw = int(e.get('total') or 0)
                cen = int(e.get('centers') or 0)
                if not name or not raw or not short: continue
                row = {'comp':label,'dist':dist,'shooter':name,'disc':short,
                       'raw':raw,'cen':cen}
                for fname, factors in FORMULAS.items():
                    row[fname] = adjust(short, raw, cen, factors)
                strings.append(row)

# ---- compute leaderboard totals: best-2 per distance per shooter per formula ----
def total_for_shooter(shooter_strings, formula_name):
    by_dist = defaultdict(list)
    for s in shooter_strings:
        by_dist[s['dist']].append(s[formula_name])
    total = 0.0
    for d, lst in by_dist.items():
        best2 = sorted(lst, reverse=True)[:2]
        total += sum(best2)
    return round(total, 3)

by_shooter = defaultdict(list)
disc_of = {}
for s in strings:
    by_shooter[s['shooter']].append(s)
    disc_of[s['shooter']] = s['disc']

leaderboard = []
for shooter, lst in by_shooter.items():
    row = {'shooter': shooter, 'disc': disc_of[shooter]}
    for fname in FORMULAS: row[fname] = total_for_shooter(lst, fname)
    row['n_dists'] = len({s['dist'] for s in lst})
    leaderboard.append(row)

# Compute ranks per formula
for fname in FORMULAS:
    sorted_by = sorted(leaderboard, key=lambda r: -r[fname])
    for i, r in enumerate(sorted_by, 1):
        r[f'rk_{fname}'] = i

# Default sort by v3 rank
leaderboard.sort(key=lambda r: r['rk_v3'])

# ---- build xlsx ----
wb = openpyxl.Workbook()
BLUE = PatternFill('solid', fgColor='1F4E78')
LITE = PatternFill('solid', fgColor='DDEBF7')
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
RED_FILL   = PatternFill('solid', fgColor='FFC7CE')
GREY_FILL  = PatternFill('solid', fgColor='F2F2F2')
WHITE_BOLD = Font(bold=True, color='FFFFFF')
HEAD = Font(bold=True)
CENTER = Alignment(horizontal='center')

def set_header(ws, cells):
    for col, val in enumerate(cells, 1):
        c = ws.cell(1, col, val)
        c.fill = BLUE; c.font = WHITE_BOLD; c.alignment = CENTER
    ws.freeze_panes = 'A2'

# --- Tab 1: Leaderboard ---
ws = wb.active
ws.title = 'Leaderboard'
header = ['Shooter','Disc',
          'v1 Total','v1 Rank',
          'v3 Total','v3 Rank',
          'Pres Total','Pres Rank',
          'Dists']
set_header(ws, header)

for i, r in enumerate(leaderboard, 2):
    ws.cell(i, 1, r['shooter'])
    ws.cell(i, 2, r['disc']).alignment = CENTER
    ws.cell(i, 3, round(r['v1'], 2))
    ws.cell(i, 4, r['rk_v1']).alignment = CENTER
    ws.cell(i, 5, round(r['v3'], 2))
    ws.cell(i, 6, r['rk_v3']).alignment = CENTER
    ws.cell(i, 7, round(r['Pres'], 2))
    ws.cell(i, 8, r['rk_Pres']).alignment = CENTER
    ws.cell(i, 9, r['n_dists']).alignment = CENTER
    if r['rk_v3'] <= 3:
        for col in range(1, len(header)+1):
            ws.cell(i, col).fill = LITE

for i, w in enumerate([26, 6, 11, 9, 11, 9, 12, 11, 7], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# --- Tab 2: Master Diffs ---
ws = wb.create_sheet('Master Diffs')
header = ['Shooter','Disc',
          'Rk v1','Rk v3','Rk Pres',
          'v1→v3 Δrank','v3→Pres Δrank','v1→Pres Δrank',
          'Total v1','Total v3','Total Pres',
          'v3 − v1','Pres − v3','Pres − v1']
set_header(ws, header)

# Order by v3 rank
for i, r in enumerate(leaderboard, 2):
    d_v1_v3 = r['rk_v1'] - r['rk_v3']      # positive = moved UP under v3
    d_v3_p  = r['rk_v3'] - r['rk_Pres']
    d_v1_p  = r['rk_v1'] - r['rk_Pres']
    def arrow(d):
        if d == 0: return '='
        return f"{'↑' if d > 0 else '↓'}{abs(d)}"
    ws.cell(i, 1, r['shooter'])
    ws.cell(i, 2, r['disc']).alignment = CENTER
    ws.cell(i, 3, r['rk_v1']).alignment = CENTER
    ws.cell(i, 4, r['rk_v3']).alignment = CENTER
    ws.cell(i, 5, r['rk_Pres']).alignment = CENTER
    for col, delta in [(6, d_v1_v3), (7, d_v3_p), (8, d_v1_p)]:
        c = ws.cell(i, col, arrow(delta))
        c.alignment = CENTER
        if delta > 0:   c.fill = GREEN_FILL
        elif delta < 0: c.fill = RED_FILL
    ws.cell(i, 9,  round(r['v1'], 2))
    ws.cell(i, 10, round(r['v3'], 2))
    ws.cell(i, 11, round(r['Pres'], 2))
    for col, dv in [(12, r['v3']-r['v1']), (13, r['Pres']-r['v3']), (14, r['Pres']-r['v1'])]:
        c = ws.cell(i, col, round(dv, 2))
        c.alignment = CENTER
        if dv > 0:   c.fill = GREEN_FILL
        elif dv < 0: c.fill = RED_FILL

for i, w in enumerate([26, 6, 7, 7, 8, 13, 14, 14, 10, 10, 11, 10, 11, 11], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

out = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_Club_Championship_3formulas.xlsx'
wb.save(out)
print(f'Wrote {out}')
print(f'  {len(leaderboard)} shooters, {len(strings)} strings, {len(FORMULAS)} formulas')
