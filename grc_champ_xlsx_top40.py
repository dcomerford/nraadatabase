"""GRC Club Championship leaderboard using V4 factors derived at TOP 40% K&Q cohort.

Same structure as grc_champ_xlsx.py — same comps, same best-2-per-distance rule,
same V4 formula structure — but factors come from the top-40%/floor-5 calibration
in v4_top40_calibration.py rather than the published V4 values.

Derived factors (vs current V4 in brackets):
  TR     1.412 (1.42)
  F-Open 1.406 (1.42)
  F-Std  1.475 (1.46)
  FTR    1.450 (1.45)
  SO     1.367 (1.40)
  SP     1.410 (1.41)
"""
import json
import urllib.request
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COMPS = [
    ('0adcba04-e8c4-47a7-9534-fdc73b52d392', 'Jan 26'),
    ('d6d4e0e5-618d-41ad-8dda-b33f7d2f1d39', 'Apr 11'),
    ('d2a4e092-a9cd-47ab-8997-593499363e17', 'May 2'),
    ('c5dee970-b934-4a0b-88ed-44f61e5986c5', 'Jun 13'),
]

# Top-40% derived factors (from v4_top40_calibration.py)
FACTORS = {'TR': 1.412, 'SP': 1.410, 'SO': 1.367, 'FO': 1.406, 'FS': 1.475, 'FTR': 1.450}
CENTRE = 0.7


def to_short(disc_name):
    if not disc_name: return None
    d = disc_name.lower()
    if 'target' in d and 'rifle' in d: return 'TR'
    if 'division open' in d:           return 'TR'
    if 'f-tr' in d or 'ftr' in d:      return 'FTR'
    if 'f-open' in d or 'fopen' in d or 'f open' in d: return 'FO'
    if 'f-standard' in d or 'f standard' in d or 'fstandard' in d: return 'FS'
    if 'production' in d:              return 'SP'
    if 'sporter open' in d:            return 'SO'
    if d.strip() == 'sporter':         return 'SP'
    return None


def adjusted(disc, total, centres):
    f = FACTORS.get(disc)
    if f is None: return None
    eq = total if disc in ('FO', 'FS', 'FTR') else total * 1.2
    return round((eq + centres * CENTRE) * f, 3)


def fetch(cid):
    return json.load(urllib.request.urlopen(
        f'https://bulletimpacts.com/api/competitions/{cid}/results'))


NAME_FIXES = {
    'Steve Gywn': 'Stephen Gwyn',
    'Nick Pato':  'Nick Patto',
}
def fix_name(n):
    return NAME_FIXES.get(n, n)


comps_data = []
all_strings = []

for cid, label in COMPS:
    d = fetch(cid)
    comp_name = d['competition']['name']
    comp_date = d['competition']['startDate']
    comp_strings = []
    for rr in d['rangeResults']:
        dist = rr['name']
        for grp in rr['results']:
            short = to_short(grp.get('disciplineName'))
            for e in grp.get('entries', []):
                name = fix_name(
                    f"{e.get('firstName','').strip()} {e.get('lastName','').strip()}".strip())
                raw = int(e.get('total') or 0)
                centres = int(e.get('centers') or 0)
                if not name or not raw or not short:
                    continue
                adj = adjusted(short, raw, centres)
                row = {
                    'comp_id': cid, 'comp_label': label,
                    'comp_name': comp_name, 'date': comp_date,
                    'dist': dist, 'shooter': name, 'disc': short,
                    'raw': raw, 'cen': centres, 'adj': adj,
                    'shot_string': e.get('shotString',''),
                }
                comp_strings.append(row)
                all_strings.append(row)
    comps_data.append({
        'cid': cid, 'label': label, 'name': comp_name, 'date': comp_date,
        'strings': comp_strings,
    })

by_shooter = defaultdict(lambda: defaultdict(list))
disc_of = {}
for s in all_strings:
    by_shooter[s['shooter']][s['dist']].append(
        (s['adj'], s['raw'], s['cen'], s['comp_label']))
    disc_of[s['shooter']] = s['disc']

leaderboard = []
for shooter, by_dist in by_shooter.items():
    total = 0.0
    breakdown = {}
    for dist, lst in by_dist.items():
        best2 = sorted(lst, key=lambda x: -x[0])[:2]
        contributed = sum(b[0] for b in best2)
        total += contributed
        breakdown[dist] = {'sum': contributed, 'count': len(lst),
                           'taken': len(best2),
                           'details': [(b[3], b[1], b[2], b[0]) for b in best2]}
    leaderboard.append({
        'shooter': shooter, 'disc': disc_of[shooter],
        'total': round(total, 3), 'breakdown': breakdown,
    })
leaderboard.sort(key=lambda r: -r['total'])

wb = openpyxl.Workbook()
BLUE = PatternFill('solid', fgColor='1F4E78')
LITE = PatternFill('solid', fgColor='DDEBF7')
WHITE_BOLD = Font(bold=True, color='FFFFFF')
HEAD = Font(bold=True)
CENTER = Alignment(horizontal='center')


def set_header(ws, cells):
    for col, val in enumerate(cells, 1):
        c = ws.cell(1, col, val)
        c.fill = BLUE; c.font = WHITE_BOLD; c.alignment = CENTER
    ws.freeze_panes = 'A2'


# --- 1. Leaderboard ---
ws = wb.active
ws.title = 'Leaderboard'
all_dists = sorted({d for s in all_strings for d in [s['dist']]},
                   key=lambda x: int(''.join(filter(str.isdigit, x)) or 0))
header = ['Rank', 'Shooter', 'Disc', 'Total Adjusted']
for d in all_dists:
    header += [f'{d} adj (best2)', f'{d} raw (best2)']
header += ['Distances Shot']
set_header(ws, header)

def fmt_raw(raw, cen):
    return f'{raw}.{cen}'

for rank, e in enumerate(leaderboard, 1):
    row = [rank, e['shooter'], e['disc'], round(e['total'], 2)]
    for d in all_dists:
        if d in e['breakdown']:
            v = e['breakdown'][d]
            row.append(round(v['sum'], 2))
            raw_pair = ' + '.join(fmt_raw(r, c) for (_, r, c, _) in v['details'])
            row.append(raw_pair)
        else:
            row.append('')
            row.append('')
    row.append(len(e['breakdown']))
    for col, val in enumerate(row, 1):
        c = ws.cell(rank+1, col, val)
        if col == 1: c.alignment = CENTER
    if rank <= 3:
        for col in range(1, len(header)+1):
            ws.cell(rank+1, col).fill = LITE

widths = [6, 26, 6, 14] + [11, 14]*len(all_dists) + [14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w


# --- 2. All Strings ---
ws = wb.create_sheet('All Strings')
header = ['Comp', 'Date', 'Distance', 'Shooter', 'Disc', 'Raw', 'Centres',
          'Adjusted', 'Shot String']
set_header(ws, header)
sorted_strings = sorted(all_strings,
    key=lambda r: (r['comp_label'], r['dist'], -r['adj']))
for i, s in enumerate(sorted_strings, 2):
    ws.cell(i, 1, s['comp_label'])
    ws.cell(i, 2, s['date'])
    ws.cell(i, 3, s['dist'])
    ws.cell(i, 4, s['shooter'])
    ws.cell(i, 5, s['disc'])
    ws.cell(i, 6, s['raw'])
    ws.cell(i, 7, s['cen'])
    ws.cell(i, 8, round(s['adj'], 2))
    ws.cell(i, 9, s['shot_string'])
for i, w in enumerate([10, 12, 9, 26, 6, 6, 8, 11, 18], 1):
    ws.column_dimensions[get_column_letter(i)].width = w


# --- 3. Factors reference sheet ---
ws = wb.create_sheet('Factors')
set_header(ws, ['Discipline', 'V4 (original)', 'V4 top-40% (this sheet)', 'Δ', 'Notes'])
v4_original = {'TR': 1.42, 'FO': 1.42, 'FS': 1.46, 'FTR': 1.45, 'SO': 1.40, 'SP': 1.41}
notes = {
    'TR':  'basically unchanged',
    'FO':  'slightly hot at top-40% — small trim',
    'FS':  'slightly cold — small bump',
    'FTR': 'unchanged',
    'SO':  'most reduced — flag sample-size caveat (489 strings only)',
    'SP':  'unchanged',
}
for i, disc in enumerate(['TR', 'FO', 'FS', 'FTR', 'SO', 'SP'], 2):
    ws.cell(i, 1, disc)
    ws.cell(i, 2, v4_original[disc])
    ws.cell(i, 3, FACTORS[disc])
    ws.cell(i, 4, round(FACTORS[disc] - v4_original[disc], 3))
    ws.cell(i, 5, notes[disc])
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 50


# --- 4..N One tab per comp ---
for comp in comps_data:
    sheet_name = comp['name'][:31]
    ws = wb.create_sheet(sheet_name)
    dists = sorted({s['dist'] for s in comp['strings']},
                   key=lambda x: int(''.join(filter(str.isdigit, x)) or 0))
    header = ['Shooter', 'Disc'] + sum([[f'{d} Raw', f'{d} Cen', f'{d} Adj']
                                        for d in dists], []) + ['Comp Adj Total']
    set_header(ws, header)

    by_sh = defaultdict(dict)
    disc_map = {}
    for s in comp['strings']:
        by_sh[s['shooter']][s['dist']] = (s['raw'], s['cen'], s['adj'])
        disc_map[s['shooter']] = s['disc']
    rows = []
    for shooter, dist_data in by_sh.items():
        comp_total = sum(v[2] for v in dist_data.values())
        rows.append((shooter, disc_map[shooter], dist_data, comp_total))
    rows.sort(key=lambda r: -r[3])
    for i, (shooter, disc, dist_data, total) in enumerate(rows, 2):
        ws.cell(i, 1, shooter); ws.cell(i, 2, disc)
        col = 3
        for d in dists:
            if d in dist_data:
                raw, cen, adj = dist_data[d]
                ws.cell(i, col, raw); ws.cell(i, col+1, cen)
                ws.cell(i, col+2, round(adj, 2))
            col += 3
        ws.cell(i, col, round(total, 2)).font = HEAD
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 6
    for i in range(3, len(header)+1):
        ws.column_dimensions[get_column_letter(i)].width = 9


out = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_Club_Championship_v4_top40.xlsx'
wb.save(out)
print(f'Wrote {out}')
print(f'  - {len(leaderboard)} shooters on leaderboard')
print(f'  - {len(all_strings)} individual range scores')
print(f'  - {len(comps_data)} competition tabs')
print()
print('Factors used (vs original V4):')
for d in ('TR', 'FO', 'FS', 'FTR', 'SO', 'SP'):
    v4 = {'TR':1.42,'FO':1.42,'FS':1.46,'FTR':1.45,'SO':1.40,'SP':1.41}[d]
    print(f'  {d:<4} {FACTORS[d]:.3f}  (was {v4:.2f}, Δ {FACTORS[d]-v4:+.3f})')
