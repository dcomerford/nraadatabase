"""Export GRC Point Shoot 13-Jun-2026 results as a multi-sheet Excel workbook.
One sheet per formula, each sorted from top MCSI to bottom.
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CSV_PATH = '/Users/dancomerford/Downloads/GRC_-_Point_Shoot_-_June_13th_2026-06-13_scores.csv'
OUT_PATH = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_13Jun2026_formula_comparison.xlsx'

YESTERDAY = {
    'Target Rifle': 1.53, 'Sporter': 1.53, 'Sporter Open': 1.47,
    'F-Standard': 1.43, 'F-Open': 1.39, 'FTR': 1.43,
}
ORIGINAL = {
    'Target Rifle': (1.62, 8.4), 'Sporter': (1.50, 12.0), 'Sporter Open': (1.50, 12.0),
    'F-Standard': (1.42, 1.8), 'F-Open': (1.42, 1.8), 'FTR': (1.42, 1.8),
}
NEW_LINEAR = {
    'Target Rifle': (1.534, 8.4), 'Sporter': (1.499, 12.0), 'Sporter Open': (1.487, 12.0),
    'F-Standard': (1.42, 1.8), 'F-Open': (1.42, 1.8), 'FTR': (1.42, 1.8),
}
PROPOSED_FACTORS = {
    'Target Rifle': 0.9796, 'Sporter': 1.0060, 'Sporter Open': 1.0060,
    'F-Standard': 1.0146, 'F-Open': 0.9708, 'FTR': 1.0160,
}
FIFTY_PT = {'Target Rifle', 'Sporter', 'Sporter Open'}


def mcsi_yesterday(raw, cs, disc):
    return round((raw + cs) * YESTERDAY[disc], 2)


def mcsi_original(raw, cs, disc):
    m, o = ORIGINAL[disc]
    return round((raw + cs) * m + o, 2)


def mcsi_new_linear(raw, cs, disc):
    m, o = NEW_LINEAR[disc]
    return round((raw + cs) * m + o, 2)


def mcsi_proposed(raw, cs, disc):
    f = PROPOSED_FACTORS[disc]
    if disc in FIFTY_PT:
        return round((raw * 1.2 + cs * 0.7) * f, 2)
    return round((raw + cs) * f, 2)


def load_rows():
    rows = []
    with open(CSV_PATH) as f:
        for r in csv.DictReader(f):
            disc = r['Discipline']
            if disc in FIFTY_PT:
                raw = int(r['Total TR score']); cs = int(r['Total TR V'])
                scoring = 'TR (50-pt)'
            else:
                raw = int(r['Total F-Class score']); cs = int(r['Total F-Class X'])
                scoring = 'F-Class (60-pt)'
            rows.append({
                'shooter': r['Shooter'],
                'club': r['Club'],
                'disc': disc,
                'scoring': scoring,
                'raw': raw,
                'cs': cs,
                'score_str': f'{raw}.{cs}',
            })
    return rows


def write_sheet(ws, title, formula_desc, rows, scorer):
    # Header rows
    ws['A1'] = 'GRC Point Shoot — 13 Jun 2026'
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')
    ws['A2'] = title
    ws['A2'].font = Font(size=12, bold=True, color='FFFFFF')
    ws['A2'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:H2')
    ws['A3'] = formula_desc
    ws['A3'].font = Font(size=10, italic=True, color='666666')
    ws.merge_cells('A3:H3')

    headers = ['Rank', 'Shooter', 'Club', 'Discipline', 'Scoring face',
               'Score (raw.V)', 'Raw', 'Centres', 'MCSI']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        c.alignment = Alignment(horizontal='center')

    # Compute MCSI, sort desc
    scored = [(r, scorer(r['raw'], r['cs'], r['disc'])) for r in rows]
    scored.sort(key=lambda x: -x[1])

    # Highlight colours
    gold = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
    silver = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    bronze = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
    thin = Border(bottom=Side(style='thin', color='CCCCCC'))

    for i, (r, mcsi) in enumerate(scored):
        row_n = 6 + i
        rank = i + 1
        values = [rank, r['shooter'], r['club'], r['disc'], r['scoring'],
                  r['score_str'], r['raw'], r['cs'], mcsi]
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row_n, column=col, value=v)
            c.border = thin
            if col in (1, 6, 7, 8, 9):
                c.alignment = Alignment(horizontal='right')
            if rank == 1: c.fill = gold
            elif rank == 2: c.fill = silver
            elif rank == 3: c.fill = bronze
        # Bold MCSI column
        ws.cell(row=row_n, column=9).font = Font(bold=True)

    # Column widths
    widths = [6, 22, 22, 16, 18, 14, 8, 10, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w


def main():
    rows = load_rows()
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sheets = [
        ('Yesterday (as used)', 'MCSI = (raw + centres) × mult     '
         'TR=1.53  FS=1.43  FO=1.39  SP=1.53  SO=1.47  (FTR=1.43 assumed)', mcsi_yesterday),
        ('Original 2014', 'MCSI = ((raw + centres) × mult) + offset     '
         'TR=1.62/8.4  F-class=1.42/1.8  Sporter=1.50/12.0', mcsi_original),
        ('New Linear (7-comp)', 'MCSI = ((raw + centres) × mult) + offset     '
         'TR=1.534/8.4  F-class=1.42/1.8  Sporter-Open=1.487/12.0  Sporter-PC=1.499/12.0', mcsi_new_linear),
        ('Proposed (22-comp)', '50-pt: (raw × 1.2 + centres × 0.7) × factor.  '
         '60-pt: (raw + centres) × factor.  '
         'TR=0.9796  Sporter=1.0060  F-Open=0.9708  F-Std=1.0146  FTR=1.0160', mcsi_proposed),
    ]
    for title, desc, scorer in sheets:
        ws = wb.create_sheet(title=title[:31])  # Excel max sheet name length
        write_sheet(ws, title, desc, rows, scorer)

    # Summary sheet
    summary = wb.create_sheet(title='Summary', index=0)
    summary['A1'] = 'GRC Point Shoot — 13 Jun 2026 — Formula Comparison'
    summary['A1'].font = Font(size=14, bold=True)
    summary.merge_cells('A1:G1')
    summary['A3'] = 'This workbook compares four MCSI formulas against yesterday\'s GRC shoot.'
    summary['A4'] = 'Each sheet shows the same 23 shooters ranked under one formula, highest MCSI first.'
    summary.merge_cells('A3:G3'); summary.merge_cells('A4:G4')

    summary['A6'] = 'Sheets:'
    summary['A6'].font = Font(bold=True)
    sheet_descs = [
        ('Yesterday (as used)', 'The formula used by GRC on the day (multiplier-only, no offset)'),
        ('Original 2014', 'The 2014 linear formula currently in the NRAA database app'),
        ('New Linear (7-comp)', 'Derived from 2024-2026 Kings/Queens (split Sporter), 7 comps'),
        ('Proposed (22-comp)', 'Different structure: (raw×1.2 + centres×0.7)×factor for 50-pt; (raw+centres)×factor for 60-pt'),
    ]
    for i, (s, d) in enumerate(sheet_descs, 7):
        summary.cell(row=i, column=1, value=s).font = Font(bold=True)
        summary.cell(row=i, column=2, value=d)
        summary.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)

    summary['A13'] = 'Top 3 under each formula:'
    summary['A13'].font = Font(bold=True)
    summary['A14'] = 'Formula'
    summary['B14'] = '1st'
    summary['C14'] = '2nd'
    summary['D14'] = '3rd'
    for c in ['A14', 'B14', 'C14', 'D14']:
        summary[c].font = Font(bold=True)
        summary[c].fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        summary[c].font = Font(bold=True, color='FFFFFF')

    for i, (title, _, scorer) in enumerate(sheets, 15):
        scored = sorted(((r, scorer(r['raw'], r['cs'], r['disc'])) for r in rows),
                        key=lambda x: -x[1])
        summary.cell(row=i, column=1, value=title).font = Font(bold=True)
        for j in range(3):
            r, m = scored[j]
            summary.cell(row=i, column=2 + j,
                         value=f'{r["shooter"]} ({r["disc"][:4]}) {m}')

    for col, w in zip('ABCD', [22, 36, 36, 36]):
        summary.column_dimensions[col].width = w

    wb.save(OUT_PATH)
    print(f'Wrote {OUT_PATH}')


if __name__ == '__main__':
    main()
