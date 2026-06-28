"""Render the GRC Club Champion 2026 leaderboard with RAW best-2 scores per
distance, PLUS a converted-score column and the MCSI total that actually drives
each ranking.

Three stacked blocks, top 10 each (30 rows), each ranked by its MCSI Grand Total:
  1. Current (Peter, Jun 13)  -> Adjusted = (Raw + Centres) x PeterFactor
  2. Last year (2025, V4)      -> standard MCSI, V4 equipment factors
  3. V5                        -> standard MCSI, V5 factors

Standard MCSI:  Adjusted = (Raw x ScoreConversion + Centres x 0.7) x Factor
  ScoreConversion = 1.2 for TR/Sporter, 1.0 for F-class.

Per-distance cells show the RAW best-2 scores (score.centres). The raw grand
total is shown as '<points>.<centres>' from those same best-2 cells. The
'Converted' column rescales everyone to a common 0-60 scale (TR/Sporter: centre
= 6 so centres ADDED; F-class: centres dropped). The 'MCSI Total' column is the
sum of the best-2 MCSI-adjusted scores per distance for that block's formula --
this is the number that determines the ranking.
"""
import openpyxl
from collections import defaultdict
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

SRC = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_MCSI_Club_Champion_Master_2026_v5_with_peter.xlsx'
OUT = '/Users/dancomerford/Desktop/claude/nraadatabase/GRC_raw_score_leaderboard_v5.png'

DISCIPLINE_LABEL = {'FO': 'F-Open', 'FS': 'F-Standard', 'FTR': 'F/TR',
                    'TR': 'TR', 'SO': 'Sporter', 'SP': 'Sporter'}
DISTS = (500, 600, 700, 800, 900)
OUT_OF_50 = {'TR', 'SO', 'SP'}   # disciplines where a centre converts to a 6

# Settings-sheet factor columns: V4=7, V3=8, V2=9, Peter's=10, V5=11
FACTOR_COL = {'V4': 7, 'V3': 8, 'V2': 9, 'Peter': 10, 'V5': 11}

# Last year's (2025) formula, reverse-engineered from allshootsummaryoct25:
#   Adjusted = mult * (Raw + Centres) + offset   (per discipline, per shoot)
# Derived exactly (R2=1.0) from the 2025 source: F-class 1.42/+0.6, TR 1.62/+2.8,
# Sporter 1.5/+4. F/TR has no 2025 sample -> treated as F-class.
Y2025 = {'FO': (1.42, 0.6), 'FS': (1.42, 0.6), 'FTR': (1.42, 0.6),
         'TR': (1.62, 2.8), 'SO': (1.5, 4.0), 'SP': (1.5, 4.0)}

# Each block: (formula label, kind, factor-set key)
#   kind 'std'   -> standard MCSI    (Raw*conv + Centres*0.7) * factor
#   kind 'peter' -> Peter's June 13  (Raw + Centres) * factor
#   kind 'y2025' -> last year's 2025 mult*(Raw + Centres) + offset
BLOCKS = [
    ('Last yr (2025)', 'y2025', None),
    ('Current (Peter)', 'peter', 'Peter'),
    ('V5', 'std', 'V5'),
]


def load():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    S = wb['Settings']
    factors = {k: {} for k in FACTOR_COL}
    conv = {}
    for r in range(2, 8):
        d = S.cell(r, 1).value
        if not d:
            continue
        conv[d] = S.cell(r, 2).value
        for key, c in FACTOR_COL.items():
            factors[key][d] = S.cell(r, c).value

    R = wb['Results']
    data = defaultdict(lambda: defaultdict(list))
    disc = {}
    for r in range(2, R.max_row + 1):
        sh = R.cell(r, 3).value
        if not sh:
            continue
        d = R.cell(r, 2).value
        di = R.cell(r, 4).value
        raw = R.cell(r, 5).value
        cen = R.cell(r, 6).value
        if raw is None or cen is None or d is None:
            continue
        data[sh][d].append((raw, int(cen)))
        disc[sh] = di
    return data, disc, factors, conv


def mcsi(raw, cen, di, kind, F, conv):
    if kind == 'peter':
        return (raw + cen) * F[di]
    if kind == 'y2025':
        m, b = Y2025[di]
        return m * (raw + cen) + b
    return (raw * conv[di] + cen * 0.7) * F[di]


def best2_raw(data, sh, d):
    """Best-1/best-2 raw shoots at a distance (by raw score, then centres)."""
    vals = sorted(data[sh].get(d, []), key=lambda x: (-x[0], -x[1]))
    return vals[0] if vals else None, vals[1] if len(vals) >= 2 else None


def fmt(rc):
    return '' if rc is None else f'{rc[0]}.{rc[1]}'


def convert(rc, di):
    """To a common 0-60 scale: TR/Sporter add centres (centre=6), F-class drop."""
    if rc is None:
        return None
    raw, cen = rc
    return raw + cen if di in OUT_OF_50 else raw


def grand_total_raw(data, sh):
    """'<points>.<centres>' from the best-2 raw cells shown in the row."""
    pts = cen = 0
    for d in DISTS:
        b1, b2 = best2_raw(data, sh, d)
        for b in (b1, b2):
            if b:
                pts += b[0]
                cen += b[1]
    return f'{pts}.{cen}'


def converted_total(data, disc, sh):
    di = disc.get(sh, '')
    t = 0
    for d in DISTS:
        b1, b2 = best2_raw(data, sh, d)
        for b in (b1, b2):
            if b:
                t += convert(b, di)
    return t


def mcsi_total(data, disc, sh, kind, F, conv):
    """Sum of best-2 MCSI-adjusted scores per distance (selected by adjusted)."""
    di = disc.get(sh, '')
    t = 0
    for d in DISTS:
        vals = sorted(data[sh].get(d, []),
                      key=lambda x: -mcsi(x[0], x[1], di, kind, F, conv))[:2]
        t += sum(mcsi(v[0], v[1], di, kind, F, conv) for v in vals)
    return round(t, 2)


def build_block(data, disc, factors, conv, label, kind, fkey, topn=10):
    F = factors[fkey] if fkey else None
    ranked = sorted((sh for sh in data if disc.get(sh)),
                    key=lambda sh: -mcsi_total(data, disc, sh, kind, F, conv))
    rows = []
    for rank, sh in enumerate(ranked[:topn], 1):
        di = disc.get(sh, '')
        cells = [sh]
        for d in DISTS:
            b1, b2 = best2_raw(data, sh, d)
            cells.append(fmt(b1))
            cells.append(fmt(b2))
        cells.append(grand_total_raw(data, sh))
        cells.append(str(converted_total(data, disc, sh)))
        cells.append(f'{mcsi_total(data, disc, sh, kind, F, conv):.2f}')
        cells.append(str(rank))
        cells.append(DISCIPLINE_LABEL.get(di, di))
        cells.append(label)
        rows.append(cells)
    return rows


def render(blocks, out_path):
    headers = ['Shooter',
               '500m\nBest 1', '500m\nBest 2', '600m\nBest 1', '600m\nBest 2',
               '700m\nBest 1', '700m\nBest 2', '800m\nBest 1', '800m\nBest 2',
               '900m\nBest 1', '900m\nBest 2',
               'Grand\nTotal (raw)', 'Converted\n(F:no cen /\nTR:+cen)',
               'MCSI\nTotal', 'Rank', 'Disc', 'Formula']
    block_fills = ['#DDEBF7', '#F2F2F2', '#E2F0D9']
    all_rows = []
    row_fill = []
    for bi, rows in enumerate(blocks):
        all_rows += rows
        row_fill += [block_fills[bi % len(block_fills)]] * len(rows)

    n_rows = len(all_rows) + 1
    n_cols = len(headers)

    fig_w = 19
    row_h = 0.40
    fig_h = row_h * n_rows + 1.4

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_xlim(0, n_cols)
    ax.set_ylim(0, n_rows + 0.5)
    ax.invert_yaxis()
    ax.axis('off')

    col_widths = [2.4] + [1.0] * 10 + [1.3, 1.1, 1.1, 0.6, 1.0, 1.2]
    total_w = sum(col_widths)
    col_widths = [w * n_cols / total_w for w in col_widths]
    col_x = [0]
    for w in col_widths:
        col_x.append(col_x[-1] + w)

    HEAD_FILL = '#1F4E78'

    ax.text(n_cols / 2, -0.15,
            'GRC Club Champion 2026 — Best 2 RAW scores per distance, '
            'with Converted + MCSI totals (top 10 per formula)',
            ha='center', va='center', fontsize=12, fontweight='bold')

    y = 0.5
    for i, h in enumerate(headers):
        ax.add_patch(mpatches.Rectangle((col_x[i], y), col_widths[i], 1.0,
                                        facecolor=HEAD_FILL, edgecolor='white'))
        ax.text(col_x[i] + col_widths[i] / 2, y + 0.5, h,
                ha='center', va='center', color='white', fontsize=7.5, fontweight='bold')

    for ri, row in enumerate(all_rows):
        y = 1.5 + ri
        fill = row_fill[ri]
        for ci, cell in enumerate(row):
            ax.add_patch(mpatches.Rectangle((col_x[ci], y), col_widths[ci], 1.0,
                                            facecolor=fill, edgecolor='#BFBFBF', linewidth=0.5))
            ha = 'left' if ci == 0 else 'center'
            xpos = col_x[ci] + 0.1 if ci == 0 else col_x[ci] + col_widths[ci] / 2
            # bold the MCSI total column (drives the ranking)
            weight = 'bold' if ci in (0, len(row) - 4) else 'normal'
            ax.text(xpos, y + 0.5, cell, ha=ha, va='center', fontsize=7.5, fontweight=weight)

    plt.tight_layout()
    plt.savefig(out_path, dpi=160, bbox_inches='tight')
    print(f'Wrote {out_path}')


if __name__ == '__main__':
    data, disc, factors, conv = load()
    blocks = [build_block(data, disc, factors, conv, label, kind, fkey)
              for (label, kind, fkey) in BLOCKS]
    render(blocks, OUT)

    for (label, kind, fkey), rows in zip(BLOCKS, blocks):
        print(f'\n{label} ranking — RawTot / Conv / MCSI:')
        print(f'{"#":>2} {"Shooter":<22} {"Disc":<11} {"RawTot":>9} {"Conv":>5} {"MCSI":>8}')
        for r in rows:
            print(f'{r[-3]:>2} {r[0]:<22} {r[-2]:<11} {r[11]:>9} {r[12]:>5} {r[13]:>8}')
