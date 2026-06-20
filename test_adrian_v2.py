"""Apply Adrian v2 (revised) factors to:
 1. Geelong June 13 — show new top-10 vs Adrian v1 and President
 2. K&Q dataset — head-to-head fairness, big-match winner distribution

V1 = Adrian's original (TR 1.45, SP 1.43, SO 1.43, FO 1.44, FS 1.46, FTR 1.46)
V2 = Adrian's new   (TR 1.415, SP 1.41, SO 1.40, FO 1.43, FS 1.46, FTR 1.50)
P  = President     (TR 1.56, SP 1.55, SO 1.46, FO 1.28, FS 1.42, FTR 1.42)
"""
from collections import defaultdict, Counter
import openpyxl, warnings
warnings.filterwarnings('ignore')
from db import get_connection

CENTRE = 0.7
V1 = {'TR':1.45,'SP':1.43,'SO':1.43,'FO':1.44,'FS':1.46,'FTR':1.46}
V2 = {'TR':1.415,'SP':1.41,'SO':1.40,'FO':1.43,'FS':1.46,'FTR':1.50}
V3 = {'TR':1.42,'SP':1.41,'SO':1.40,'FO':1.43,'FS':1.46,'FTR':1.46}
P  = {'TR':1.56,'SP':1.55,'SO':1.46,'FO':1.28,'FS':1.42,'FTR':1.42}

def adj(raw, centres, disc, factors):
    is_f = disc in ('FO','FS','FTR')
    eq = raw if is_f else raw * 1.2
    return round((eq + centres * CENTRE) * factors[disc], 3)

# ---------- 1. GEELONG comparison ----------
wb = openpyxl.load_workbook(
    '/Users/dancomerford/Downloads/geelong_rifle_club_adjusted_scores_june_26 (1).xlsx',
    data_only=True)
ws = wb['Raw Printout Order']
rows = []
for r in range(4, ws.max_row+1):
    name = ws.cell(r,2).value
    if not name: continue
    disc = ws.cell(r,3).value
    raw = int(ws.cell(r,5).value or 0)
    cen = int(ws.cell(r,6).value or 0)
    if not raw: continue
    rows.append({
        'name': name, 'disc': disc, 'raw': raw, 'cen': cen,
        'v1': adj(raw, cen, disc, V1),
        'v2': adj(raw, cen, disc, V2),
        'v3': adj(raw, cen, disc, V3),
        'p' : adj(raw, cen, disc, P),
    })

def rk(metric):
    s = sorted(rows, key=lambda e: -e[metric])
    return {e['name']: i+1 for i, e in enumerate(s)}
r_v1, r_v2, r_p = rk('v1'), rk('v2'), rk('p')

print('='*100)
print('GEELONG 13-JUN — top 20 by ADRIAN v2 (new factors)')
print('='*100)
print(f"{'rk(v2)':<7}{'shooter':<20}{'disc':<5}{'raw':>5}{'cen':>4}"
      f"{'v2':>10}{'v1':>10}{'pres':>10}{'rk(v1)':>9}{'rk(P)':>9}")
print('-'*100)
top = sorted(rows, key=lambda e: -e['v2'])[:20]
for i, e in enumerate(top, 1):
    print(f"{i:<7}{e['name'][:18]:<20}{e['disc']:<5}{e['raw']:>5}{e['cen']:>4}"
          f"{e['v2']:>10.2f}{e['v1']:>10.2f}{e['p']:>10.2f}"
          f"{r_v1[e['name']]:>9}{r_p[e['name']]:>9}")

# Composition of top 10
print('\n=== Top 10 composition by formula ===')
for label, ranker in [('v1', r_v1), ('v2', r_v2), ('Pres', r_p)]:
    top10 = sorted(rows, key=lambda e: -e[{'v1':'v1','v2':'v2','Pres':'p'}[label]])[:10]
    counts = Counter(e['disc'] for e in top10)
    print(f"  {label:<5} top-10: " + ', '.join(f'{d}×{n}' for d,n in counts.most_common()))

# Place movement v1 → v2
print('\n=== Biggest movers from Adrian v1 to v2 ===')
moves = [(r_v1[e['name']] - r_v2[e['name']], e) for e in rows]
moves.sort(key=lambda x: -abs(x[0]))
for m, e in moves[:8]:
    if m == 0: continue
    arrow = '↑' if m > 0 else '↓'
    print(f"  {e['name'][:18]:<20}{e['disc']:<5}{r_v1[e['name']]:>3} → {r_v2[e['name']]:<3} {arrow}{abs(m)}")

# Theoretical ceilings
print('\n=== Theoretical max per discipline (perfect with all 10 centres, 10 shots) ===')
print(f"  {'disc':<5}{'v1':>8}{'v2':>8}{'P':>8}{'v2 vs v1':>11}")
for d in ('TR','SO','SP','FO','FS','FTR'):
    is_f = d in ('FO','FS','FTR')
    raw, max_c = (60, 10) if is_f else (50, 10)
    eq = raw if is_f else raw * 1.2
    base = eq + max_c * CENTRE
    print(f"  {d:<5}{base*V1[d]:>8.2f}{base*V2[d]:>8.2f}{base*P[d]:>8.2f}"
          f"{base*V2[d]-base*V1[d]:>+10.2f}")

# ---------- 2. K&Q broader test ----------
print('\n' + '='*100)
print('K&Q 2024+ TEST — head-to-head & big-match winner distribution under v2')
print('='*100)

cur = get_connection().cursor()
cur.execute("""
    SELECT s.competition_id, s.match_number, s.match_name, s.discipline,
           s.score, s.shots_raw, s.target_max
    FROM strings s JOIN competitions c ON c.competition_id=s.competition_id
    JOIN states st ON st.state_id=c.state_id
    WHERE s.is_kings_queens=TRUE AND s.score IS NOT NULL AND c.year>=2024
      AND st.code = ANY(%s) AND s.discipline IS NOT NULL
      AND (SELECT COUNT(*) FROM shots WHERE string_id=s.string_id)=10
""", (['NSWRA','VRA','QRA','NQRA','SARA','WARA','NRAA'],))

# Map discipline → short code
def to_short(disc, tmax):
    if disc in ('TR-A','TR-B','TR-C','Division Open') and tmax==5: return 'TR'
    if disc == 'F-Open' and tmax==6: return 'FO'
    if disc.startswith('F-Std') and tmax==6: return 'FS'
    if disc == 'FTR' and tmax==6: return 'FTR'
    if disc == 'Sporter-Open' and tmax==5: return 'SO'
    if disc == 'Sporter-PC' and tmax==5: return 'SP'
    if disc == 'Sporter-Combined' and tmax==5: return 'SP'  # fold into PC
    return None

enriched = []
for (comp, mnum, mname, disc, score, shots, tmax) in cur.fetchall():
    short = to_short(disc, tmax)
    if not short: continue
    raw_int = int(float(score))
    max_raw = 60 if short in ('FO','FS','FTR') else 50
    if raw_int > max_raw: continue
    centres = sum(1 for ch in (shots or '') if ch == ('V' if tmax==5 else 'X'))
    enriched.append({
        'comp':comp,'mnum':mnum,'disc':short,
        'v1':adj(raw_int, centres, short, V1),
        'v2':adj(raw_int, centres, short, V2),
        'v3':adj(raw_int, centres, short, V3),
        'p' :adj(raw_int, centres, short, P),
    })

by_match = defaultdict(list)
for e in enriched: by_match[(e['comp'], e['mnum'])].append(e)

# Big-match winners
print('\n=== Big match winner distribution (≥20 shooters) ===')
big = [es for es in by_match.values() if len(es) >= 20]
print(f"matches: {len(big)}")
print(f"  {'disc':<5}{'v1':>10}{'v2':>10}{'pres':>10}")
for which in ('v1','v2','p'):
    pass
cnt_v1 = Counter(); cnt_v2 = Counter(); cnt_p = Counter()
for es in big:
    cnt_v1[max(es,key=lambda e:e['v1'])['disc']] += 1
    cnt_v2[max(es,key=lambda e:e['v2'])['disc']] += 1
    cnt_p [max(es,key=lambda e:e['p' ])['disc']] += 1
total = len(big)
for d in ('TR','SO','SP','FO','FS','FTR'):
    print(f"  {d:<5}{cnt_v1[d]:>4d}({100*cnt_v1[d]/total:>3.0f}%)"
          f"{cnt_v2[d]:>5d}({100*cnt_v2[d]/total:>3.0f}%)"
          f"{cnt_p[d]:>5d}({100*cnt_p[d]/total:>3.0f}%)")

# Head-to-head pairwise
def h2h(metric):
    out = Counter()
    for es in by_match.values():
        if len(es) < 5: continue
        tops = {}
        for e in es:
            cur = tops.get(e['disc'])
            if cur is None or e[metric] > cur[metric]:
                tops[e['disc']] = e
        present = list(tops.keys())
        for i in range(len(present)):
            for j in range(i+1, len(present)):
                a, b = present[i], present[j]
                if tops[a][metric] > tops[b][metric]: out[(a,b)] += 1
                elif tops[b][metric] > tops[a][metric]: out[(b,a)] += 1
    return out

h_v1 = h2h('v1'); h_v2 = h2h('v2'); h_p = h2h('p')
print('\n=== Head-to-head: A beats B in top-of-cat-per-match ===')
cats = ('TR','FO','FS','FTR','SO','SP')
print(f"  {'pair':<14}{'v1':>10}{'v2':>10}{'pres':>10}")
for a in cats:
    for b in cats:
        if a==b: continue
        t1=h_v1[(a,b)]+h_v1[(b,a)]; t2=h_v2[(a,b)]+h_v2[(b,a)]; tp=h_p[(a,b)]+h_p[(b,a)]
        if max(t1,t2,tp) < 10: continue
        p1=100*h_v1[(a,b)]/t1 if t1 else 0
        p2=100*h_v2[(a,b)]/t2 if t2 else 0
        pp=100*h_p [(a,b)]/tp if tp else 0
        def fmt(x): return f'{x:>6.1f}%{" ⚠" if x>=70 or x<=30 else "  "}'
        print(f"  {a:<3}>{b:<10}{fmt(p1)}{fmt(p2)}{fmt(pp)}")
